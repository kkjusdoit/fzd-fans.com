#!/usr/bin/env python3
"""Extract match data from Excel to JSON for the 战绩查询系统 page."""

import json
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path.home() / "Downloads" / "樊振东职业生涯战绩查询系统-260601.xlsx"
DATA_DIR = Path(__file__).resolve().parent.parent / "src" / "data"


def extract_matches(wb):
    ws = wb["⭐FANtastic⭐"]
    records = []
    cur_period = ""
    cur_tournament = ""
    cur_level = ""
    cur_date_range = ""
    cur_event_detail = ""
    cur_placement = ""

    for row_idx in range(3, ws.max_row + 1):
        c1 = ws.cell(row=row_idx, column=1).value
        c2 = ws.cell(row=row_idx, column=2).value
        c3 = ws.cell(row=row_idx, column=3).value
        c4 = ws.cell(row=row_idx, column=4).value
        c5 = ws.cell(row=row_idx, column=5).value
        c6 = ws.cell(row=row_idx, column=6).value
        c8 = ws.cell(row=row_idx, column=8).value
        c9 = ws.cell(row=row_idx, column=9).value
        c11 = ws.cell(row=row_idx, column=11).value
        c12 = ws.cell(row=row_idx, column=12).value
        c13 = ws.cell(row=row_idx, column=13).value
        c15 = ws.cell(row=row_idx, column=15).value
        c17 = ws.cell(row=row_idx, column=17).value
        c18 = ws.cell(row=row_idx, column=18).value
        c19 = ws.cell(row=row_idx, column=19).value

        if c1 is not None: cur_period = str(c1).strip()
        if c2 is not None: cur_tournament = str(c2).strip()
        if c3 is not None: cur_level = str(c3).strip()
        if c4 is not None: cur_date_range = str(c4).strip()
        if c5 is not None: cur_event_detail = str(c5).strip()
        if c6 is not None: cur_placement = str(c6).strip()

        if c17 is None:
            continue

        date_str = str(c11).strip() if c11 else ""
        sort_key = "0000.00.00"
        if date_str and date_str != "——":
            parts = date_str.split(".")
            if len(parts) == 3:
                sort_key = f"{parts[0].zfill(4)}.{parts[1].zfill(2)}.{parts[2].zfill(2)}"

        event = str(c8).strip() if c8 else ""
        partner = str(c9).strip() if c9 else ""
        if partner == "——":
            partner = None

        records.append({
            "period": cur_period,
            "tournament": cur_tournament,
            "level": cur_level,
            "event": event,
            "date": date_str if date_str else None,
            "dateSort": sort_key,
            "round": str(c12).strip() if c12 else "",
            "partner": partner,
            "assoc": str(c13).strip() if c13 else "",
            "opponent": str(c15).strip() if c15 else "",
            "result": str(c17).strip(),
            "score": str(c18).strip() if c18 else "",
            "games": str(c19).strip() if c19 else "",
        })

    records.sort(key=lambda r: r["dateSort"] if r["dateSort"] else "", reverse=True)

    out = DATA_DIR / "matches.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=None, separators=(",", ":"))
    print(f"Matches: {len(records)} records → {out}")


def extract_rankings(wb):
    ws = wb["看板数据"]
    rankings = []
    for row_idx in range(2, ws.max_row + 1):
        period = ws.cell(row=row_idx, column=1).value
        rank = ws.cell(row=row_idx, column=2).value
        points = ws.cell(row=row_idx, column=3).value
        if period and rank:
            rankings.append({
                "date": str(period)[:10],
                "rank": int(rank),
                "points": int(points) if points else 0,
            })
    rankings.sort(key=lambda r: r["date"])

    out = DATA_DIR / "rankings.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(rankings, f, ensure_ascii=False, indent=None, separators=(",", ":"))
    print(f"Rankings: {len(rankings)} entries → {out}")


def extract_yearly(wb):
    ws = wb["看板数据"]
    yearly = []
    for row_idx in range(2, ws.max_row + 1):
        year = ws.cell(row=row_idx, column=6).value
        event = ws.cell(row=row_idx, column=7).value
        total = ws.cell(row=row_idx, column=8).value
        wins = ws.cell(row=row_idx, column=9).value
        rate = ws.cell(row=row_idx, column=10).value
        ext_total = ws.cell(row=row_idx, column=11).value
        ext_wins = ws.cell(row=row_idx, column=12).value
        ext_rate = ws.cell(row=row_idx, column=13).value
        if year and event:
            r = float(rate) if rate and rate != "-" else None
            er = float(ext_rate) if ext_rate and ext_rate != "-" else None
            yearly.append({
                "year": str(int(year)),
                "event": str(event),
                "total": int(total) if total else 0,
                "wins": int(wins) if wins else 0,
                "rate": round(r * 100, 1) if r is not None else None,
                "extTotal": int(ext_total) if ext_total else 0,
                "extWins": int(ext_wins) if ext_wins else 0,
                "extRate": round(er * 100, 1) if er is not None else None,
            })

    out = DATA_DIR / "yearly.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(yearly, f, ensure_ascii=False, indent=None, separators=(",", ":"))
    print(f"Yearly: {len(yearly)} entries → {out}")


def extract():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    extract_matches(wb)
    extract_rankings(wb)
    extract_yearly(wb)


if __name__ == "__main__":
    extract()
